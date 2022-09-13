import openpyxl

def getRowCount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetname)
    return (sheet.max_row)

def getColumnCount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetname)
    return (sheet.max_column)

def readData(file,sheetname,rownum, coloumnno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetname)
    return sheet.cell(row=rownum, column=coloumnno).value

def writeData(file,sheetname,rownum, coloumnno,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetname)
    sheet.cell(row=rownum, column=coloumnno).value=data
    workbook.save(file)